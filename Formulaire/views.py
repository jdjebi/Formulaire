from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django import forms
from Formulaire.models import Patient

import openpyxl

class PatientCreateFrom(forms.ModelForm):
    class Meta:
        model = Patient
        fields = (
            'nom_patient',
            'vulnerabilite',
            'groupe_risque',
            'score_risque'
        )

def patients_list(request):

    patients = Patient.objects.all()

    return render(request,"Formulaire/patients/patients_list.html",{
        "patients": patients
    })

def patients_show(request):
    return render(request,"Formulaire/patients/patients_list.html")

def patients_create(request):

    if request.method == "POST":

        form = PatientCreateFrom(request.POST)

        print(request.POST)
        #print(form.errors)
        
        if form.is_valid():
            p_data = {}

            date_creation = form.data.get('date_creation')
            date_naissance = form.data.get('date_naissance')

            # Les conditions ci-dessous permettent d'éviter les conflits de format de date avec django au cas les champs date sont vides 
            if date_creation == '':
                date_creation = None
            
            if date_naissance == '':
                date_naissance = None

            p_data["date_creation"] = date_creation
            p_data["date_creation_str"] = form.data.get('date_creation')
            p_data["nom_patient"] = form.data.get('nom_patient')
            p_data["sexe"] = form.data.get('sexe')
            p_data["date_naissance"] = date_naissance
            p_data["date_naissance_str"] = form.data.get('date_naissance')
            p_data["poids"] = form.data.get('poids')
            p_data["taille"] = form.data.get('taille')

            p_data["pa_systotique"] = form.data.get('pa_systotique')
            p_data["pa_diastolique"] = form.data.get('pa_diastolique')
            p_data["pam"] = form.data.get('pam')
            p_data["rythme_cardiaque_pa"] = form.data.get('rythme_cardiaque_pa')
            p_data["hta"] = form.data.get('hta')

            p_data["oxymetrie"] = form.data.get('oxymetrie')
            p_data["rythme_cardiaque_oxy"] = form.data.get('rythme_cardiaque_oxy')

            p_data["temperature"] = form.data.get('temperature')

            p_data["imc"] = form.data.get('imc')
            p_data["interpretation_imc"] = form.data.get('interpretation_imc')

            p_data["gyclémie_1"] = form.data.get('glycemie_1')
            p_data["gyclémie_2"] = form.data.get('glycemie_2')

            p_data["diabete"] = form.data.get('diabete')
            p_data["cardiopathie"] = form.data.get('cardiopathie')
            p_data["ecg"] = form.data.get('ecg')
            p_data["interpretation_ecg"] = form.data.get('interpretation_ecg')

            p_data["vulnerabilite"] = form.data.get('vulnerabilite')
            p_data["groupe_risque"] = form.data.get('groupe_risque')
            p_data["score_risque"] = form.data.get('score_risque')

            create_patient(p_data)

            messages.success(request, 'Patient créé !')


    return render(request,"Formulaire/patients/patients_create.html")

def patients_import(request):

    if request.method == "POST" and request.FILES:
    
        patients_file_excel = request.FILES["fichier_patients"]

        wb = openpyxl.load_workbook(patients_file_excel)

        excel_data = []

        worksheet = wb.active
        
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                value = str(cell.value).strip()
                if value == 'None' or value == "":
                    value = None
                row_data.append(value)
            print(row_data[5])
            excel_data.append(row_data)

        print(excel_data[0])
            
        excel_data = excel_data[1:] # On commence à partir de la deuxième ligne, le première contient les colonnes

        for data in excel_data:
            # Règle d'acceptation: La liste de données à l'indice 2 c'est à le nom du patient ne doit pas être à 'None'
            if data[2] != None:
                p_data = {}
                p_data["date_creation"] = None
                p_data["date_creation_str"] = data[1]
                p_data["nom_patient"] = data[2]
                p_data["sexe"] = data[3]
                p_data["date_naissance"] = data[4]
                p_data["date_naissance_str"] = data[4]
                p_data["poids"] = None
                p_data["taille"] = data[6]

                p_data["pa_systotique"] = data[7]
                p_data["pa_diastolique"] = data[8]
                p_data["pam"] = data[9]
                p_data["rythme_cardiaque_pa"] = data[10]
                p_data["hta"] = data[11]

                p_data["oxymetrie"] = data[12]
                p_data["rythme_cardiaque_oxy"] = data[13]

                p_data["temperature"] = data[14]

                p_data["imc"] = data[15]
                p_data["interpretation_imc"] = data[16]

                p_data["gyclémie_1"] = None
                p_data["gyclémie_2"] = None

                p_data["diabete"] = data[17]
                p_data["cardiopathie"] = data[18]
                p_data["ecg"] = data[19]
                p_data["interpretation_ecg"] = data[20]

                p_data["vulnerabilite"] = data[21]
                p_data["groupe_risque"] = data[22]
                p_data["score_risque"] = data[23]

                create_patient(p_data)
        
        messages.success(request, 'Importation réussie')

    return redirect("patients.list")

def patients_export(request):

    patients = Patient.objects.all()

    wb = openpyxl.Workbook()

    ws = wb.active

    titles = ['Individu_ID', 'Date de Création', 'Nom Patient', 'Sexe', 'Date de Naissance', 'Poids', 'Tailles', 'PA Systolique', 'PA Diastolique', 'PAM', 'PA Rythme cardiaque', 'HTA', 'Oxymetrie %', 'Oxymetrie Rythme cardiaque', 'Températures', 'IMC', 'Interprétation IMC', 'Diabète', 'Cardiopathie', 'ECG', 'Interprétation ECG', 'Classifier la vulnérabilité', "CLASSIFIER LE GROUPE DE RISQUE D'EXPOSITION", 'Score de risque','Glycémie (mg/dl)','Glycémie (mmol/L)']

    ws.append(titles)

    for patient in patients:
        data = []
        
        data = [
            patient.id,
            str(patient.date_creation),
            patient.nom_patient,
            patient.sexe,
            str(patient.date_naissance),
            patient.poids,
            patient.taille,
            patient.pa_systotique,
            patient.pa_diastolique,
            patient.pam,
            patient.rythme_cardiaque_pa,
            patient.hta,
            patient.oxymetrie,
            patient.rythme_cardiaque_oxy,
            patient.temperature,
            patient.imc,
            patient.interpretation_imc,
            patient.diabete,
            patient.cardiopathie,
            patient.ecg,
            patient.interpretation_ecg,
            patient.vulnerabilite,
            patient.groupe_risque,
            patient.score_risque,
            patient.gyclémie_1,
            patient.gyclémie_2,
        ]
        
        ws.append(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

    wb.save(response)
       
    return response

def create_patient(data):
    patient = Patient(
        nom_patient=data["nom_patient"],
        date_creation=data["date_creation"],
        date_creation_str=data["date_creation_str"],
        sexe=data["sexe"],
        date_naissance=data["date_naissance"],
        date_naissance_str=data["date_naissance_str"],
        poids=data["poids"],
        taille=data["taille"],

        pa_systotique=data["pa_systotique"],
        pa_diastolique=data["pa_diastolique"],
        pam=data["pam"],
        rythme_cardiaque_pa=data["rythme_cardiaque_pa"],
        hta=data["hta"],

        oxymetrie=data["oxymetrie"],
        rythme_cardiaque_oxy=data["rythme_cardiaque_oxy"],

        temperature=data["temperature"],

        imc=data["imc"],
        interpretation_imc=data["interpretation_imc"],

        gyclémie_1=data["gyclémie_1"],
        gyclémie_2=data["gyclémie_2"],
        diabete=data["diabete"],
        cardiopathie=data["cardiopathie"],
        ecg=data["ecg"],
        interpretation_ecg=data["interpretation_ecg"],

        vulnerabilite=data["vulnerabilite"],
        groupe_risque=data["groupe_risque"],
        score_risque=data["score_risque"],
    )

    patient.save()

    print(patient.pk)